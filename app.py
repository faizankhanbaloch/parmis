from __future__ import annotations

import sqlite3
from datetime import datetime, timezone
from typing import Optional

from fastapi import FastAPI, Form, Request
from fastapi.responses import HTMLResponse, Response
from fastapi.staticfiles import StaticFiles


import os
import threading
from datetime import datetime
from openpyxl import Workbook, load_workbook

EXCEL_PATH = "leads.xlsx"
_excel_lock = threading.Lock()

EXCEL_HEADERS = [
    "created_at",
    "name",
    "phone",
    "email",
    "state",
    "service",
    "message",
    "ip",
]

def append_lead_to_excel(lead: dict) -> None:
    """
    Appends a lead to leads.xlsx (creates file + headers if missing).
    Thread-safe for concurrent requests.
    """
    with _excel_lock:
        if not os.path.exists(EXCEL_PATH):
            wb = Workbook()
            ws = wb.active
            ws.title = "Leads"
            ws.append(EXCEL_HEADERS)
            wb.save(EXCEL_PATH)

        wb = load_workbook(EXCEL_PATH)
        ws = wb["Leads"] if "Leads" in wb.sheetnames else wb.active

        row = [lead.get(h, "") for h in EXCEL_HEADERS]
        ws.append(row)

        wb.save(EXCEL_PATH)




APP_NAME = "PARMIS Painting Services"
TAGLINE = "Prep-First Craftsmanship. Premium finishes. Zero mess."

PORTFOLIO_ITEMS = [
    {"img": "/static/portfolio/portfolio-1.png", "title": "Coastal exterior repaint", "subtitle": "Weather-ready, crisp trims"},
    {"img": "/static/portfolio/portfolio-2.png", "title": "Modern commercial frontage", "subtitle": "Durable finish, minimal downtime"},
    {"img": "/static/portfolio/portfolio-3.png", "title": "Living room refresh", "subtitle": "Smooth walls + clean cut lines"},
    {"img": "/static/portfolio/portfolio-4.png", "title": "Feature wall + staircase", "subtitle": "Colour balance and detail work"},
    {"img": "/static/portfolio/portfolio-5.png", "title": "Surface prep & patching", "subtitle": "Repairs, sanding, paint-ready"},
    {"img": "/static/portfolio/portfolio-6.png", "title": "Premium finish kitchen", "subtitle": "Even coats, long-lasting sheen"},
]

LEARN_PREVIEW = {
  "residential": "/static/residential.png",
  "commercial": "/static/commercial.png",
  "prep": "/static/surface-prep.png",
  "premium-finish": "/static/premium-finish.png",
}
SERVICE_AREA = "Australia"
DB_PATH = "leads.sqlite3"

app = FastAPI(title=APP_NAME)
app.mount("/static", StaticFiles(directory="static"), name="static")


def init_db() -> None:
    with sqlite3.connect(DB_PATH) as con:
        con.execute(
            """
            CREATE TABLE IF NOT EXISTS leads (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              created_at TEXT NOT NULL,
              name TEXT NOT NULL,
              phone TEXT NOT NULL,
              email TEXT,
              suburb TEXT NOT NULL,
              service TEXT NOT NULL,
              message TEXT NOT NULL,
              page TEXT
            )
            """
        )
        con.commit()


@app.on_event("startup")
def _startup() -> None:
    init_db()


# -----------------------------
# “Generated images” (SVG) endpoints
# -----------------------------
def svg_brandmark() -> str:
    return """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 120 120" role="img" aria-label="Brandmark">
  <defs>
    <linearGradient id="g" x1="0" y1="0" x2="1" y2="1">
      <stop offset="0" stop-color="#7dd3fc"/>
      <stop offset=".55" stop-color="#c4b5fd"/>
      <stop offset="1" stop-color="#34d399"/>
    </linearGradient>
    <filter id="b" x="-40%" y="-40%" width="180%" height="180%">
      <feGaussianBlur stdDeviation="4"/>
    </filter>
  </defs>
  <rect x="10" y="10" width="100" height="100" rx="30" fill="rgba(255,255,255,.05)" stroke="rgba(255,255,255,.14)"/>
  <path d="M30 70c12 0 18-12 30-12s18 12 30 12" fill="none" stroke="url(#g)" stroke-width="10" stroke-linecap="round"/>
  <path d="M28 46c6 0 9-5 14-5s8 5 14 5 9-5 14-5 8 5 14 5" fill="none" stroke="rgba(255,255,255,.55)" stroke-width="7" stroke-linecap="round"/>
  <circle cx="36" cy="86" r="4" fill="#7dd3fc" filter="url(#b)" opacity=".9"/>
  <circle cx="60" cy="90" r="4" fill="#c4b5fd" filter="url(#b)" opacity=".9"/>
  <circle cx="84" cy="86" r="4" fill="#34d399" filter="url(#b)" opacity=".9"/>
</svg>"""


def svg_art(seed: int, label: str) -> str:
    # “photo-like” abstract finish boards (generated)
    # (not icons; feels like modern portfolio thumbnails)
    a = ["#7dd3fc", "#c4b5fd", "#34d399", "#fb7185", "#fbbf24", "#60a5fa"]
    c1 = a[(seed * 3) % len(a)]
    c2 = a[(seed * 5 + 1) % len(a)]
    c3 = a[(seed * 7 + 2) % len(a)]
    return f"""<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 1400 900">
  <defs>
    <linearGradient id="bg" x1="0" y1="0" x2="1" y2="1">
      <stop offset="0" stop-color="{c1}" stop-opacity=".35"/>
      <stop offset=".55" stop-color="{c2}" stop-opacity=".20"/>
      <stop offset="1" stop-color="{c3}" stop-opacity=".18"/>
    </linearGradient>
    <linearGradient id="s" x1="0" y1="0" x2="1" y2="0">
      <stop offset="0" stop-color="{c1}"/>
      <stop offset="1" stop-color="{c2}"/>
    </linearGradient>
    <filter id="blur" x="-30%" y="-30%" width="160%" height="160%">
      <feGaussianBlur stdDeviation="22"/>
    </filter>
    <filter id="grain">
      <feTurbulence type="fractalNoise" baseFrequency=".8" numOctaves="3" stitchTiles="stitch" />
      <feColorMatrix type="matrix" values="1 0 0 0 0  0 1 0 0 0  0 0 1 0 0  0 0 0 .18 0"/>
    </filter>
  </defs>

  <rect width="1400" height="900" fill="rgba(255,255,255,.03)"/>
  <rect width="1400" height="900" fill="url(#bg)"/>

  <circle cx="260" cy="240" r="240" fill="{c1}" opacity=".16" filter="url(#blur)"/>
  <circle cx="1120" cy="280" r="300" fill="{c2}" opacity=".12" filter="url(#blur)"/>
  <circle cx="760" cy="760" r="360" fill="{c3}" opacity=".10" filter="url(#blur)"/>

  <path d="M160 580 C360 460 520 520 700 420 C900 320 1080 360 1240 260"
        fill="none" stroke="url(#s)" stroke-width="34" stroke-linecap="round" opacity=".55"/>
  <path d="M150 660 C400 590 560 640 770 570 C980 500 1100 540 1260 470"
        fill="none" stroke="rgba(255,255,255,.55)" stroke-width="18" stroke-linecap="round" opacity=".28"/>

  <g opacity=".88">
    <rect x="380" y="390" width="640" height="380" rx="26" fill="rgba(0,0,0,.18)" stroke="rgba(255,255,255,.14)"/>
    <path d="M420 710 H980" stroke="rgba(255,255,255,.35)" stroke-width="10" stroke-linecap="round"/>
    <path d="M420 650 H860" stroke="rgba(255,255,255,.28)" stroke-width="10" stroke-linecap="round"/>
    <path d="M420 590 H940" stroke="rgba(255,255,255,.22)" stroke-width="10" stroke-linecap="round"/>
  </g>

  <rect width="1400" height="900" filter="url(#grain)" opacity=".55"/>

  <text x="72" y="120" fill="rgba(255,255,255,.92)"
        font-family="ui-sans-serif, system-ui, -apple-system, Segoe UI, Roboto, Helvetica, Arial"
        font-weight="900" font-size="54">{label}</text>
  <text x="72" y="168" fill="rgba(226,232,240,.78)"
        font-family="ui-sans-serif, system-ui, -apple-system, Segoe UI, Roboto, Helvetica, Arial"
        font-weight="700" font-size="22">Generated portfolio artwork</text>
</svg>"""


@app.get("/img/brand.svg")
def img_brand():
    return Response(svg_brandmark(), media_type="image/svg+xml")

def svg_figma_bg() -> str:
    # Matches the Figma hero panel: soft top-left highlight + light blue wash + subtle vignette
    return """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 1600 900" preserveAspectRatio="none">
  <defs>
    <linearGradient id="wash" x1="0" y1="0" x2="0" y2="1">
      <stop offset="0" stop-color="#f6f9ff"/>
      <stop offset="1" stop-color="#eef4ff"/>
    </linearGradient>

    <radialGradient id="spot" cx="18%" cy="20%" r="70%">
      <stop offset="0" stop-color="#ffffff" stop-opacity="0.95"/>
      <stop offset="0.40" stop-color="#dfeaff" stop-opacity="0.55"/>
      <stop offset="1" stop-color="#dfeaff" stop-opacity="0"/>
    </radialGradient>

    <radialGradient id="vignette" cx="50%" cy="35%" r="85%">
      <stop offset="0" stop-color="#000000" stop-opacity="0"/>
      <stop offset="1" stop-color="#000000" stop-opacity="0.08"/>
    </radialGradient>

    <filter id="grain">
      <feTurbulence type="fractalNoise" baseFrequency=".9" numOctaves="3" stitchTiles="stitch"/>
      <feColorMatrix type="matrix"
        values="1 0 0 0 0  0 1 0 0 0  0 0 1 0 0  0 0 0 .08 0"/>
    </filter>
  </defs>

  <rect width="1600" height="900" fill="url(#wash)"/>
  <rect width="1600" height="900" fill="url(#spot)"/>
  <rect width="1600" height="900" fill="url(#vignette)"/>
  <rect width="1600" height="900" filter="url(#grain)" opacity=".45"/>
</svg>"""


@app.get("/img/figma-bg.svg")
def img_figma_bg():
    return Response(svg_figma_bg(), media_type="image/svg+xml")
@app.get("/img/art/{seed}.svg")
def img_art(seed: int):
    label = [
        "Interior Finish — Clean Edges",
        "Exterior Refresh — Weather Ready",
        "Commercial — Durable Coats",
        "Surface Prep — Smooth Base",
        "Premium Finish — Even Coverage",
        "Detail Work — Trims & Doors",
        "Retail Space — After Hours",
        "Apartment Repaint — Low Odour",
    ][seed % 8]
    return Response(svg_art(seed, label), media_type="image/svg+xml")


# -----------------------------
# HTML shell
# -----------------------------
def page(title: str, path: str, content: str) -> str:
    year = datetime.now().year

    def nav(href: str) -> str:
        return "navItem active" if href == path else "navItem"


    return f"""<!doctype html>
<html lang="en-AU">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>{title} · {APP_NAME}</title>
  <meta name="description" content="{TAGLINE}"/>
  <meta name="theme-color" content="#070b14"/>

  <script src="https://cdn.tailwindcss.com"></script>

  <!-- GSAP animations -->
  <script src="https://cdnjs.cloudflare.com/ajax/libs/gsap/3.12.5/gsap.min.js"></script>
  <script src="https://cdnjs.cloudflare.com/ajax/libs/gsap/3.12.5/ScrollTrigger.min.js"></script>

  <!-- HTMX for quote submit -->
  <script src="https://unpkg.com/htmx.org@1.9.12"></script>

  <style>
    :root {{
      --bg:#070b14;
      --panel:rgba(255,255,255,.06);
      --panel2:rgba(255,255,255,.04);
      --line:rgba(255,255,255,.12);
      --line2:rgba(255,255,255,.18);
      --shadow:0 24px 70px rgba(0,0,0,.45);
      --r2:26px;
    }}
    html{{scroll-behavior:smooth}}
    body{{background:var(--bg)}}
    .card{{ background: rgba(255,255,255,.75); border: 1px solid rgba(15,23,42,.08); box-shadow: 0 20px 60px rgba(2,6,23,.10); }}
    .card2{{ background: rgba(255,255,255,.60); border: 1px solid rgba(15,23,42,.08); }}

    .heroWrap{{
  width: 100%;
  background: linear-gradient(180deg, rgba(245,242,255,1) 0%, rgba(240,248,245,1) 100%);
  border-bottom: 1px solid rgba(15,23,42,.08);
}}
.heroInner{{
  max-width: 1200px;
  margin: 0 auto;
  padding: 56px 28px;
}}
@media(min-width:768px){{
  .heroInner{{ padding: 82px 36px; }}
}}
.heroImageCard{{
  border-radius: 18px;
  overflow: hidden;
  box-shadow: 0 24px 60px rgba(2,6,23,.18);
  border: 1px solid rgba(15,23,42,.10);
  background: #fff;
}}
.heroImageCard img{{
  width: 100%;
  height: 100%;
  display:block;
  object-fit: cover;
}}
.statPill{{
  position:absolute;
  left:-26px;
  bottom: 26px;
  background:#fff;
  border: 1px solid rgba(15,23,42,.10);
  border-radius: 14px;
  box-shadow: 0 16px 40px rgba(2,6,23,.12);
  padding: 14px 18px;
  min-width: 170px;
}}
.navItem{{
  padding: 10px 14px;
  border-radius: 999px;
  font-weight: 800;
  font-size: 14px;
  color: #0f172a;
  opacity: .75;
  transition: all .18s ease;
}}
.navItem:hover{{ opacity: 1; background: rgba(15,23,42,.06); }}

.navItem.active{{
  opacity: 1;
  background: rgba(15,23,42,.10);
  box-shadow: inset 0 0 0 1px rgba(15,23,42,.10);
}}
.statCenter{{
  left: 50%;
  top: 50%;
  bottom: auto;
  transform: translate(-50%, -50%);
}}




    .btn{{border-radius:999px; padding:12px 16px; font-weight:900; font-size:14px; display:inline-flex; align-items:center; justify-content:center; gap:10px}}
    .btn-primary{{ background:#0b1220; color:#fff; }}
    .btn-primary:hover{{opacity:.92}}
    .btn-ghost{{
  border:1px solid rgba(15,23,42,.12);
  background: rgba(255,255,255,.75);
  color: #0f172a;
}}
.btn-ghost:hover{{
  background: rgba(255,255,255,.95);
  border-color: rgba(15,23,42,.18);
}}
.noticeSuccess{{
  background: #dcfce7;              /* green-100 */
  border: 1px solid #86efac;        /* green-300 */
  color: #064e3b;                   /* emerald-900 */
}}
.noticeSuccess .title{{ color:#064e3b; font-weight:900; }}
.noticeSuccess .sub{{ color:#065f46; opacity:1; }}
.field{{
  width:100%;
  border-radius:16px;
  border:1px solid rgba(15,23,42,.12);
  background: rgba(255,255,255,.85);
  padding:12px 14px;
  color:#0f172a;
  outline:none;
}}
.field::placeholder{{ color: rgba(71,85,105,.75); }}
.field:focus{{ border-color: rgba(15,23,42,.22); background: rgba(255,255,255,.98); }}


    /* premium animated background layer */
    .bgGlow{{
  position: fixed;
  inset: 0;
  z-index: -10;
  pointer-events: none;
  background-size: cover;
  background-position: center;
  background-repeat: no-repeat;
}}

body{{ background: #ffffff; }}

    /* hero canvas */
    #paintCanvas{{border-radius:28px; border:1px solid rgba(255,255,255,.12); background:rgba(0,0,0,.25)}}

    /* 3D tilt cards */
    .tilt{{transform-style:preserve-3d; perspective:900px}}
    .tiltInner{{transition: transform .12s ease; will-change: transform}}
    .ink{{
      position:absolute; inset:-40%;
      background:
        radial-gradient(500px 280px at 20% 20%, rgba(125,211,252,.18), transparent 60%),
        radial-gradient(540px 300px at 80% 30%, rgba(196,181,253,.16), transparent 60%),    
        radial-gradient(520px 320px at 45% 90%, rgba(52,211,153,.12), transparent 60%);
      filter: blur(28px);
      opacity:.0;
      transition: opacity .25s ease;
      pointer-events:none;
    }}
    .tilt:hover .ink{{opacity:.95}}

    /* modal */
    .modal{{position:fixed; inset:0; display:none; align-items:center; justify-content:center; padding:18px; z-index:80; background:rgba(0,0,0,.65); backdrop-filter:blur(10px)}}
    .modal.open{{display:flex}}
    .modalCard{{width:min(980px,100%); border-radius:28px; border:1px solid rgba(255,255,255,.12); background:rgba(2,6,23,.76); overflow:hidden; box-shadow:var(--shadow)}}

    /* masonry-like */
    .masonry{{columns:1; column-gap:14px}}
    @media(min-width:768px){{.masonry{{columns:3}}}}
    .mTile{{break-inside:avoid; margin-bottom:14px}}

    /* reveal on scroll */
    .reveal{{opacity:0; transform: translateY(18px)}}
  </style>
</head>

<body class="min-h-screen text-slate-900">
  <div class="bgGlow"></div>

    <header class="sticky top-0 z-50 border-b border-slate-200/60 bg-slate-200/90 backdrop-blur">
    <div class="mx-auto flex max-w-7xl items-center justify-between px-6 py-4">
        
        <!-- Left: Logo + name -->
        <a href="/" class="flex items-center gap-4">
        <img src="/static/logo.png"
            class="h-14 w-14 object-contain"
            alt={APP_NAME}/>
        <div class="leading-tight">
            <div class="text-lg font-extrabold tracking-tight text-slate-900">{APP_NAME}</div>
            <div class="text-sm text-slate-600">{TAGLINE}</div>
        </div>
        </a>

        <!-- Center: Nav -->
        <nav class="hidden md:flex items-center gap-2 rounded-full bg-white/60 px-2 py-1 shadow-sm ring-1 ring-slate-200">
        <a class="{nav('/')}" href="/">Home</a>
        <a class="{nav('/services')}" href="/services">Services</a>
        <a class="{nav('/portfolio')}" href="/portfolio">Portfolio</a>
        <a class="{nav('/contact')}" href="/contact">Contact</a>
        </nav>

        <!-- Right: CTA -->
        <a href="/contact" class="btn btn-primary px-5 py-3 rounded-full font-extrabold">
        Free Quote
        </a>

    </div>
    </header>




  

  <main>{content}</main>

  <footer class="border-t border-white/10">
    <div class="w-full px-6 md:px-10 py-10">
      <div class="grid gap-6 md:grid-cols-3">
        <div class="card2 p-6">
          <div class="flex items-center gap-3">
            <img src="/static/logo.png" class="h-9 w-9" alt=""/>
            <div class="text-base font-extrabold">{APP_NAME}</div>
          </div>
          <p class="mt-3 text-sm text-slate-600">Prep-first painting. Premium materials. Clean, respectful service.</p>
          <p class="mt-2 text-sm text-slate-600">Serving {SERVICE_AREA}.</p>
        </div>

        <div class="card2 p-6">
          <div class="text-sm font-bold">Explore</div>
          <div class="mt-3 grid gap-2 text-sm text-slate-600">
            <a class="hover:underline" href="/services">Services</a>
            <a class="hover:underline" href="/portfolio">Portfolio</a>
            <a class="hover:underline" href="/contact">Get a quote</a>
          </div>
        </div>

        <div class="card2 p-6">
          <div class="text-sm font-bold">Fast quote</div>
          <p class="mt-3 text-sm text-slate-600">Send a request and we’ll respond as soon as possible.</p>
          <div class="mt-4 flex gap-2">
            <a class="btn btn-primary" href="/contact">Request quote</a>
            <a class="btn btn-ghost" href="/portfolio">View work</a>
          </div>
        </div>
      </div>

      <div class="mt-8 flex flex-col gap-2 border-t border-white/10 pt-6 text-xs text-slate-600 md:flex-row md:justify-between">
        <div>© {year} {APP_NAME}. All rights reserved.</div>
        <div class="flex gap-4">
          <a class="hover:underline" href="/robots.txt">Robots</a>
          <a class="hover:underline" href="/sitemap.xml">Sitemap</a>
        </div>
      </div>
    </div>
  </footer>

  <script>
    // ---- GSAP reveal-on-scroll ----
    gsap.registerPlugin(ScrollTrigger);
    document.querySelectorAll(".reveal").forEach((el) => {{
      gsap.to(el, {{
        opacity: 1,
        y: 0,
        duration: 0.8,
        ease: "power2.out",
        scrollTrigger: {{
          trigger: el,
          start: "top 86%"
        }}
      }});
    }});

    // ---- 3D tilt cards ----
    function initTilt() {{
      document.querySelectorAll(".tilt").forEach((card) => {{
        const inner = card.querySelector(".tiltInner");
        if (!inner) return;

        function onMove(e) {{
          const r = card.getBoundingClientRect();
          const px = (e.clientX - r.left) / r.width;
          const py = (e.clientY - r.top) / r.height;
          const rx = (py - 0.5) * -10;
          const ry = (px - 0.5) * 14;
          inner.style.transform = "rotateX(" + rx + "deg) rotateY(" + ry + "deg) translateZ(0)";
        }}
        function onLeave() {{
          inner.style.transform = "rotateX(0deg) rotateY(0deg)";
        }}
        card.addEventListener("mousemove", onMove);
        card.addEventListener("mouseleave", onLeave);
      }});
    }}
    initTilt();

    // ---- Portfolio modal ----
    const modal = document.getElementById("modal");
    const modalClose = document.getElementById("modalClose");
    const modalImg = document.getElementById("modalImg");
    const modalTitle = document.getElementById("modalTitle");

    function openModal(src, title) {{
      modalImg.src = src;
      modalTitle.textContent = title;
      modal.classList.add("open");
      modal.setAttribute("aria-hidden","false");
    }}
    function closeModal() {{
      modal.classList.remove("open");
      modal.setAttribute("aria-hidden","true");
    }}
    document.addEventListener("click", (e) => {{
      const t = e.target.closest("[data-open]");
      if (t) openModal(t.dataset.open, t.dataset.title || "Project");
      if (e.target === modal || e.target === modalClose) closeModal();
    }});
    document.addEventListener("keydown", (e)=>{{ if(e.key==="Escape") closeModal(); }});

    // ---- Animated “paint flow” canvas (lightweight, no libs) ----
    (function() {{
      const c = document.getElementById("paintCanvas");
      if (!c) return;
      const ctx = c.getContext("2d");
      let w=0,h=0,t=0;

      function resize() {{
        const dpr = Math.min(2, window.devicePixelRatio || 1);
        const rect = c.getBoundingClientRect();
        w = Math.floor(rect.width * dpr);
        h = Math.floor(rect.height * dpr);
        c.width = w; c.height = h;
        ctx.setTransform(1,0,0,1,0,0);
      }}
      window.addEventListener("resize", resize);
      resize();

      function blob(x,y,r,alpha, col) {{
        const g = ctx.createRadialGradient(x,y,0,x,y,r);
        g.addColorStop(0, col.replace("ALPHA", String(alpha)));
        g.addColorStop(1, col.replace("ALPHA", "0"));
        ctx.fillStyle = g;
        ctx.beginPath(); ctx.arc(x,y,r,0,Math.PI*2); ctx.fill();
      }}

      function loop() {{
        t += 0.012;
        ctx.clearRect(0,0,w,h);

        // base wash
        ctx.fillStyle = "rgba(0,0,0,0.18)";
        ctx.fillRect(0,0,w,h);

        // moving blobs (paint)
        const cols = [
          "rgba(125,211,252,ALPHA)",
          "rgba(196,181,253,ALPHA)",
          "rgba(52,211,153,ALPHA)"
        ];

        for (let i=0;i<14;i++) {{
          const k = i * 0.45;
          const x = (0.5 + 0.42*Math.sin(t*1.2 + k)) * w;
          const y = (0.5 + 0.36*Math.cos(t*1.05 + k*1.3)) * h;
          const r = (0.14 + 0.08*Math.sin(t + k)) * Math.min(w,h);
          blob(x,y,r,0.22, cols[i%cols.length]);
        }}

        // subtle streaks
        ctx.globalAlpha = 0.12;
        ctx.strokeStyle = "white";
        ctx.lineWidth = Math.max(1, Math.min(6, w/260));
        ctx.beginPath();
        for(let x=0; x<w; x+=w/14) {{
          const yy = h*0.55 + Math.sin(t*1.8 + x*0.01)*h*0.08;
          ctx.moveTo(x, yy);
          ctx.quadraticCurveTo(x+w*0.08, yy-h*0.06, x+w*0.16, yy+h*0.02);
        }}
        ctx.stroke();
        ctx.globalAlpha = 1;

        requestAnimationFrame(loop);
      }}
      loop();
    }})();
  </script>
</body>
</html>
"""


# -----------------------------
# Routes
# -----------------------------
@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    content = f"""
<section class="heroWrap">
  <div class="heroInner">
    <div class="grid gap-10 md:grid-cols-2 md:items-center">
      <div>
        <h1 class="text-5xl md:text-6xl font-black tracking-tight text-slate-900 leading-[1.05]">
          Transform Your Home<br/>with Professional Painting
        </h1>
        <p class="mt-5 text-lg text-slate-600 max-w-xl">
          Quality craftsmanship, premium materials, and clean respectful service — built to last.
        </p>

        <div class="mt-7 flex flex-wrap gap-3">
          <a href="/contact" class="btn btn-primary">Get Free Estimate</a>
          <a href="/portfolio" class="btn btn-ghost">View Our Work</a>
        </div>

        <ul class="mt-8 space-y-3 text-sm text-slate-700">
          <li class="flex items-center gap-2">
            <span class="inline-block h-5 w-5 rounded-full bg-emerald-100 border border-emerald-200"></span>
            Licensed & Insured
          </li>
          <li class="flex items-center gap-2">
            <span class="inline-block h-5 w-5 rounded-full bg-emerald-100 border border-emerald-200"></span>
            15+ Years Experience
          </li>
          <li class="flex items-center gap-2">
            <span class="inline-block h-5 w-5 rounded-full bg-emerald-100 border border-emerald-200"></span>
            100% Satisfaction Guarantee
          </li>
        </ul>
      </div>

      <div class="relative">
        <div class="heroImageCard">
          <img src="/static/hero-collage.png" alt="Residential and commercial painting projects"/>
        </div>

        <div class="statPill statCenter">
            <div class="text-2xl font-black text-slate-900">500+</div>
            <div class="text-sm text-slate-600">Happy Customers</div>
        </div>

      </div>
    </div>
  </div>
</section>

<section class="w-full px-6 md:px-10 pb-14">
  <div class="reveal card p-7 md:p-10">
    <div class="flex flex-col gap-2 md:flex-row md:items-end md:justify-between">
      <div>
        <h2 class="text-2xl font-black tracking-tight">Services</h2>    
      </div>
      <a href="/services" class="btn btn-ghost">See all services</a>
    </div>

    <div class="mt-7 grid gap-5 md:grid-cols-2">
      {service_card("Residential painting", "Interior + exterior. Walls, ceilings, doors, trims. Colour guidance available.", "Fast turnarounds with clean protection and tidy handover.","/learn/residential")}
      {service_card("Commercial painting", "Offices, retail, strata maintenance. Scheduling to minimise disruption.", "Durable coatings with professional-grade materials and consistent finish.","/learn/commercial")}
      {service_card("Surface preparation", "Repairs, patching, crack fill, sanding, old paint removal.", "Prep-first approach for adhesion and long-lasting results.","/learn/prep")}
      {service_card("Premium finish", "High-quality paints, smooth coats, detailed edge work.", "Straight lines, even coverage and a final walkthrough.", "/learn/premium-finish")}
    </div>
  </div>
</section>

<section class="w-full px-6 md:px-10 pb-14">
  <div class="reveal card p-7 md:p-10">
    <h2 class="text-2xl font-black tracking-tight">Work in action (video)</h2>

    <div class="mt-6 grid gap-5 md:grid-cols-3">
      {video_card("Exterior repaint time-lapse", "")}
      {video_card("Interior painting time-lapse", "")}
      {video_card("Commercial spray painting time-lapse", "")}
    </div>
  </div>
</section>

<section class="w-full px-6 md:px-10 pb-14">
  <div class="reveal card p-7 md:p-10">
    <div class="flex flex-col gap-2 md:flex-row md:items-end md:justify-between">
      <div>
        <h2 class="text-2xl font-black tracking-tight">Portfolio</h2>
        <p class="mt-2 text-sm text-slate-600">Generated “photo-like” boards. Click to open a preview modal.</p>
      </div>
      <a href="/portfolio" class="btn btn-ghost">Open full portfolio</a>
    </div>

    <div class="mt-7 masonry">
      {"".join(portfolio_tile(p["img"], p["title"], p["subtitle"]) for p in PORTFOLIO_ITEMS)}
    </div>
  </div>
</section>

<section class="w-full px-6 md:px-10 pb-16">
  <div class="reveal card p-7 md:p-10">
    <div class="grid gap-8 md:grid-cols-2 md:items-start">
      <div>
        <h2 class="text-2xl font-black tracking-tight">Get a free quote</h2>
        <p class="mt-3 text-sm text-slate-600">
          Tell us what you need and we’ll respond as soon as possible.
        </p>
        <div class="mt-5 grid gap-3">
          <div class="card2 p-4">
            <div class="text-sm font-extrabold">Fast response</div>
            <div class="mt-1 text-sm text-slate-600">Clear estimate and timeline.</div>
          </div>
          <div class="card2 p-4">
            <div class="text-sm font-extrabold">Clean site</div>
            <div class="mt-1 text-sm text-slate-600">Protection, tidy work, walkthrough.</div>
          </div>
        </div>
      </div>

      <div class="card2 p-6">
        <form class="grid gap-3" hx-post="/api/quote" hx-target="#quoteResult" hx-swap="innerHTML">
          <input type="hidden" name="page" value="/" />
          <div class="grid gap-3 md:grid-cols-2">
            <input name="name" class="field" placeholder="Name" required/>
            <input name="phone" class="field" placeholder="Phone" required/>
          </div>
          <input name="email" class="field" placeholder="Email (optional)"/>
          <div class="grid gap-3 md:grid-cols-2">
            <input name="suburb" class="field" placeholder="Suburb" required/>
            <select name="service" class="field" required>
              <option value="" selected>Select service</option>
              <option>Residential painting</option>
              <option>Commercial painting</option>
              <option>Surface preparation</option>
              <option>Premium finish / repaint</option>
            </select>
          </div>
          <textarea name="message" class="field min-h-[120px]" placeholder="Job details (areas, repairs, timeframe)…" required></textarea>
          <button class="btn btn-primary w-full" type="submit">Send request</button>
          <div id="quoteResult"></div>
          <div class="text-xs text-slate-600">Saved locally to SQLite when deployed.</div>
        </form>
      </div>
    </div>
  </div>
</section>

<div id="modal" class="modal" aria-hidden="true">
  <div class="modalCard">
    <div class="flex items-center justify-between border-b border-white/10 px-5 py-4">
      <div>
        <div id="modalTitle" class="text-sm font-extrabold">Preview</div>
        <div class="text-xs text-slate-600">Generated artwork</div>
      </div>
      <button id="modalClose" class="btn btn-ghost py-2 px-3 text-xs">Close</button>
    </div>
    <div class="bg-black/5 p-4">
      <img id="modalImg" class="w-full max-h-[70vh] object-contain bg-white/5" src="/static/portfolio/portfolio-1.png" alt="Preview"/>
    </div>
    <div class="grid gap-4 px-5 py-5 md:grid-cols-3">
      <div class="card2 p-4">
        <div class="text-xs font-extrabold">Prep</div>
        <div class="mt-1 text-sm text-slate-600">Repairs, sanding and clean base for adhesion.</div>
      </div>
      <div class="card2 p-4">
        <div class="text-xs font-extrabold">Finish</div>
        <div class="mt-1 text-sm text-slate-600">Even coats and sharp edge work.</div>
      </div>
      <div class="card2 p-4">
        <div class="text-xs font-extrabold">Care</div>
        <div class="mt-1 text-sm text-slate-600">Protection and tidy handover.</div>
      </div>
    </div>
  </div>
</div>
"""
    return HTMLResponse(page("Home", "/", content))

def service_card(title: str, line1: str, line2: str, href: str) -> str:
    return f"""
<div class="tilt relative overflow-hidden card2 p-6">
  <div class="ink"></div>
  <div class="tiltInner relative">
    <div class="text-lg font-black">{title}</div>
    <p class="mt-2 text-sm text-slate-600">{line1}</p>
    <p class="mt-3 text-sm text-slate-600">{line2}</p>
    <a href="{href}" class="mt-5 inline-flex items-center gap-2 text-sm font-extrabold text-slate-900 hover:opacity-80">
      Learn more <span class="opacity-70" aria-hidden="true">→</span>
    </a>
  </div>
</div>
"""


def portfolio_tile(img: str, title: str, subtitle: str) -> str:
    return f"""
<div class="mTile">
  <button class="w-full text-left card2 overflow-hidden hover:opacity-[.98]"
          data-open="{img}"
          data-title="{title}">
    <img src="{img}" alt="{title}" class="w-full h-[240px] object-cover"/>
    <div class="p-4">
      <div class="text-sm font-extrabold">{title}</div>
      <div class="mt-1 text-xs text-slate-600">{subtitle}</div>
    </div>
  </button>
</div>
"""


def video_card(title: str, youtube_id: str) -> str:
    return f"""
<div class="card2 overflow-hidden">
  <div class="aspect-video bg-black/30">
    <iframe class="h-full w-full"
      src="https://www.youtube.com/embed/{youtube_id}?rel=0&modestbranding=1"
      title="{title}"
      frameborder="0"
      allow="accelerometer; autoplay; clipboard-write; encrypted-media; gyroscope; picture-in-picture; web-share"
      allowfullscreen></iframe>
  </div>
  <div class="p-4">
    <div class="text-sm font-extrabold">{title}</div>
    <div class="mt-1 text-xs text-slate-600">Embedded example video</div>
  </div>
</div>
"""


@app.get("/services", response_class=HTMLResponse)
def services(request: Request):
    content = f"""
<section class="w-full px-6 md:px-10 pb-14 pt-10 md:pt-14">
  <div class="reveal max-w-2xl">
    <h1 class="text-4xl font-black tracking-tight">Services</h1>
    <p class="mt-3 text-lg text-slate-600">
      Prep-first workmanship and premium finishing for residential and commercial spaces.
    </p>
  </div>

  <div class="mt-8 grid gap-6 md:grid-cols-2">
    <div class="reveal card p-7">
      <div class="text-xl font-black">Residential painting</div>
      <ul class="mt-4 list-disc space-y-2 pl-5 text-sm text-slate-700">
        <li>Interior & exterior house painting</li>
        <li>Walls, ceilings, doors, and trims</li>
        <li>Apartments, townhouses, and villas</li>
        <li>Colour consultation available</li>
        <li>Clean, tidy, respectful service</li>
      </ul>
    </div>

    <div class="reveal card p-7">
      <div class="text-xl font-black">Commercial painting</div>
      <ul class="mt-4 list-disc space-y-2 pl-5 text-sm text-slate-700">
        <li>Offices, shops, and retail spaces</li>
        <li>Strata & property maintenance painting</li>
        <li>Flexible scheduling to minimise disruption</li>
        <li>Durable, professional-grade finishes</li>
      </ul>
    </div>

    <div class="reveal card p-7">
      <div class="text-xl font-black">Surface preparation</div>
      <ul class="mt-4 list-disc space-y-2 pl-5 text-sm text-slate-700">
        <li>Wall repairs and patching</li>
        <li>Crack filling and plaster touch-ups</li>
        <li>Sanding and smoothing surfaces</li>
        <li>Removal of old, peeling paint</li>
      </ul>
    </div>

    <div class="reveal card p-7">
      <div class="text-xl font-black">Premium finish</div>
      <ul class="mt-4 list-disc space-y-2 pl-5 text-sm text-slate-700">
        <li>High-quality paints and materials</li>
        <li>Smooth, even coats with long-lasting results</li>
        <li>Detail work on edges and corners</li>
      </ul>
    </div>
  </div>

  <div class="reveal mt-10 card p-8">
    <div class="flex flex-col gap-4 md:flex-row md:items-center md:justify-between">
      <div>
        <div class="text-lg font-black">Ready to book?</div>
        <div class="mt-1 text-sm text-slate-600">Send a quote request — we’ll reply as soon as possible.</div>
      </div>
      <a class="btn btn-primary" href="/contact">Get a free quote</a>
    </div>
  </div>
</section>
"""
    return HTMLResponse(page("Services", "/services", content))



def learn_layout(
    slug: str,
    title: str,
    subtitle: str,
    highlights: list[str],
    what_included: list[str],
    process: list[str],
    faqs: list[tuple[str, str]],
    preview: str
) -> str:
    # Uses generated artwork as a visual header (swap with real photos later)
    faq_html = "".join(
        f"""
        <div class="card2 p-5">
          <div class="text-sm font-extrabold">{q}</div>
          <div class="mt-2 text-sm text-slate-600">{a}</div>
        </div>
        """
        for q, a in faqs
    )
    return f"""
<section class="heroWrap">
  <div class="heroInner">
    <div class="grid gap-10 md:grid-cols-2 md:items-center">
      <div>
        <div class="text-sm font-extrabold text-slate-600">Service</div>
        <h1 class="mt-2 text-4xl md:text-5xl font-black tracking-tight text-slate-900 leading-[1.06]">
          {title}
        </h1>
        <p class="mt-4 text-lg text-slate-600 max-w-xl">{subtitle}</p>

        <div class="mt-6 grid gap-3">
          {"".join(f'<div class="card2 p-4"><div class="text-sm font-extrabold">{h}</div></div>' for h in highlights)}
        </div>

        <div class="mt-7 flex flex-wrap gap-3">
          <a href="/contact" class="btn btn-primary">Get a free quote</a>
          <a href="/services#{slug}" class="btn btn-ghost">Back to services</a>
        </div>
      </div>

      <div class="relative">
        <div class="heroImageCard">
          <img src="{preview}" alt="{title} preview"/>
        </div>
      </div>
    </div>
  </div>
</section>

<section class="w-full px-6 md:px-10 pb-14 pt-10">
  <div class="grid gap-6 md:grid-cols-2">
    <div class="reveal card p-7">
      <div class="text-xl font-black">What’s included</div>
      <ul class="mt-4 list-disc space-y-2 pl-5 text-sm text-slate-700">
        {"".join(f"<li>{x}</li>" for x in what_included)}
      </ul>
    </div>

    <div class="reveal card p-7">
      <div class="text-xl font-black">Our process</div>
      <ol class="mt-4 list-decimal space-y-2 pl-5 text-sm text-slate-700">
        {"".join(f"<li>{x}</li>" for x in process)}
      </ol>
      <div class="mt-5 text-sm text-slate-600">Want a tailored plan? We’ll walk the site and quote accurately.</div>
    </div>
  </div>

  <div class="reveal mt-8 card p-7">
    <div class="text-xl font-black">FAQs</div>
    <div class="mt-5 grid gap-4 md:grid-cols-2">
      {faq_html}
    </div>
  </div>

  <div class="reveal mt-10 card p-8">
    <div class="flex flex-col gap-4 md:flex-row md:items-center md:justify-between">
      <div>
        <div class="text-lg font-black">Ready to get started?</div>
        <div class="mt-1 text-sm text-slate-600">Send a quote request — we’ll reply as soon as possible.</div>
      </div>
      <a class="btn btn-primary" href="/contact">Request a free quote</a>
    </div>
  </div>
</section>
"""

@app.get("/learn/residential", response_class=HTMLResponse)
def learn_residential(request: Request):
    content = learn_layout(
        slug="residential",
        title="Residential painting",
        subtitle="Interior and exterior painting with clean preparation, crisp edges, and a tidy handover.",
        highlights=["Low-mess setup", "Colour guidance available", "Respectful, tidy worksite"],
        what_included=[
            "Walls, ceilings, doors and trims",
            "Apartments, townhouses, villas and houses",
            "Surface prep, patching and crack filling as needed",
            "Premium paints for a smooth, even finish",
            "Final walkthrough and touch-ups",
        ],
        process=[
            "Protect floors and furniture, mask and cover surfaces",
            "Patch, sand, and prime where required",
            "Cut-in edges and detail areas",
            "Apply 2+ coats for even coverage",
            "Clean-up and final walkthrough",
        ],
        faqs=[
            ("How long does a typical job take?", "Most interiors take 1–3 days depending on size and prep. We’ll confirm a timeline after an onsite look."),
            ("Do you help with colour selection?", "Yes — we can guide you on finishes, sheen levels, and colour direction that suits your space and lighting."),
            ("Will you protect my floors and furniture?", "Always. We use drop sheets, masking, and careful setup to keep things clean."),
            ("Do I need to move everything?", "We can work around furniture where possible. We’ll tell you exactly what needs moving before we start."),
        ],
        preview = LEARN_PREVIEW.get("residential", "/static/residential.png")
    )
    
    return HTMLResponse(page("Residential painting", "/learn/residential", content))


@app.get("/learn/commercial", response_class=HTMLResponse)
def learn_commercial(request: Request):
    content = learn_layout(
        slug="commercial",
        title="Commercial painting",
        subtitle="Durable coatings for offices, retail, and strata — scheduled to minimise disruption.",
        highlights=["After-hours options", "Durable, pro-grade finishes", "Strata & maintenance ready"],
        what_included=[
            "Offices, shops, retail and common areas",
            "Strata and property maintenance painting",
            "Low-odour options where needed",
            "Durable coatings for high-traffic areas",
            "Clear staging plan and site communication",
        ],
        process=[
            "Plan staging around trading/tenants",
            "Protect surfaces and set up clean work zones",
            "Repair, sand and prime as required",
            "Apply specified coating system for durability",
            "Handover with checklists and touch-ups",
        ],
        faqs=[
            ("Can you work after hours?", "Yes — we can schedule evenings/weekends to keep disruption low."),
            ("Do you provide maintenance plans?", "We can suggest a maintenance schedule for high-traffic areas and strata common spaces."),
            ("Can you match existing colours?", "Yes. We can colour-match or work from existing paint codes."),
            ("Are your coatings commercial grade?", "We use quality systems suited to the surface and traffic level."),
        ],
        preview = LEARN_PREVIEW.get("commercial", "/static/commercial.png")
    )
    return HTMLResponse(page("Commercial painting", "/learn/commercial", content))


@app.get("/learn/prep", response_class=HTMLResponse)
def learn_prep(request: Request):
    content = learn_layout(
        slug="prep",
        title="Surface preparation",
        subtitle="The foundation of a premium finish: repairs, patching, sanding, and clean priming.",
        highlights=["Smooth base", "Better adhesion", "Longer-lasting finish"],
        what_included=[
            "Wall repairs and patching",
            "Crack filling and plaster touch-ups",
            "Sanding and smoothing for an even surface",
            "Removal of peeling/flaking paint",
            "Priming for adhesion and stain blocking",
        ],
        process=[
            "Inspect surfaces and identify failures (cracks, peeling, stains)",
            "Scrape loose material and feather edges",
            "Patch and sand to a smooth finish",
            "Spot-prime repairs and problem areas",
            "Final dust-down and readiness check",
        ],
        faqs=[
            ("Why does prep matter so much?", "Prep determines adhesion and how smooth the final finish looks — it’s the difference between ‘good’ and ‘premium’."),
            ("Can you fix cracks and holes?", "Yes — we patch, fill, and sand before painting."),
            ("Do you remove peeling paint?", "We remove loose paint and stabilise edges so new coats won’t fail."),
            ("Will the wall look perfectly smooth?", "We aim for a clean, even result. For ‘level 5’ finishes we can discuss additional skim/finishing options."),
        ],
        preview = LEARN_PREVIEW.get("prep", "/static/prep.png")
    )
    return HTMLResponse(page("Surface preparation", "/learn/prep", content))


@app.get("/learn/premium-finish", response_class=HTMLResponse)
def learn_premium_finish(request: Request):
    content = learn_layout(
        slug="finish",
        title="Premium finish",
        subtitle="High-quality materials, sharp edge work, and consistent coverage for a polished result.",
        highlights=["Premium paints", "Crisp cut-ins", "Even, durable coats"],
        what_included=[
            "High-quality paints and materials",
            "Smooth, even coats with long-lasting results",
            "Detail work on edges and corners",
            "Trim/door finishing as needed",
            "Final walkthrough and touch-ups",
        ],
        process=[
            "Confirm sheen/finish and colour direction",
            "Cut-in edges and detail areas cleanly",
            "Apply 2+ coats for consistency",
            "Check for holidays/patchiness under lighting",
            "Final touch-ups and walkthrough",
        ],
        faqs=[
            ("What makes a finish ‘premium’?", "Even coverage, clean edges, correct sheen, and careful lighting checks — plus quality materials."),
            ("Do you use premium paint brands?", "We use quality paints suited to the job and can recommend options based on durability and washability."),
            ("Can you do feature walls?", "Yes — feature colours and different sheen levels are no problem."),
            ("Do you guarantee your work?", "We stand by our workmanship and will address any reasonable issues after completion."),
        ],
        preview = LEARN_PREVIEW.get("finish", "/static/finish.png")
    )
    return HTMLResponse(page("Premium finish", "/learn/premium-finish", content))



@app.get("/portfolio", response_class=HTMLResponse)
def portfolio(request: Request):
    content = """
<section class="w-full px-6 md:px-10 pb-14 pt-10 md:pt-14">
  <div class="reveal max-w-2xl">
    <h1 class="text-4xl font-black tracking-tight">Portfolio</h1>
    <p class="mt-3 text-lg text-slate-600">
      Generated previews (swap for real photos later). Click any tile for a modal preview.
    </p>
  </div>

  <div class="mt-8 masonry reveal">
    """ + "".join(portfolio_tile(p["img"], p["title"], p["subtitle"]) for p in PORTFOLIO_ITEMS) + """
  </div>
</section>

<div id="modal" class="modal" aria-hidden="true">
  <div class="modalCard">
    <div class="flex items-center justify-between border-b border-white/10 px-5 py-4">
      <div>
        <div id="modalTitle" class="text-sm font-extrabold">Preview</div>
        <div class="text-xs text-slate-600">Generated artwork</div>
      </div>
      <button id="modalClose" class="btn btn-ghost py-2 px-3 text-xs">Close</button>
    </div>
    <div class="bg-black/5 p-4">
      <img id="modalImg" class="w-full max-h-[70vh] object-contain bg-white/5" src="/static/portfolio/portfolio-1.png" alt="Preview"/>
    </div>
  </div>
</div>
"""
    return HTMLResponse(page("Portfolio", "/portfolio", content))


@app.get("/contact", response_class=HTMLResponse)
def contact(request: Request):
    content = """
<section class="w-full px-6 md:px-10 pb-14 pt-10 md:pt-14">
  <div class="grid gap-10 md:grid-cols-2 md:items-start">
    <div class="reveal">
      <h1 class="text-4xl font-black tracking-tight">Get a free quote</h1>
      <p class="mt-4 text-lg text-slate-600">
        The more detail you include, the more accurate the quote.
      </p>

      <div class="mt-6 grid gap-4">
        <div class="card2 p-5">
          <div class="text-sm font-extrabold">Include</div>
          <ul class="mt-2 list-disc space-y-2 pl-5 text-sm text-slate-600">
            <li>Interior/exterior and areas</li>
            <li>Repairs required (cracks, peeling)</li>
            <li>Preferred timeframe</li>
            <li>Colour direction (optional)</li>
          </ul>
        </div>
      </div>
    </div>

    <div class="reveal card p-7">
      <div class="text-sm font-extrabold">Quote request</div>
      <form class="mt-4 grid gap-3" hx-post="/api/quote" hx-target="#quoteResult" hx-swap="innerHTML">
        <input type="hidden" name="page" value="/contact" />
        <div class="grid gap-3 md:grid-cols-2">
          <input name="name" class="field" placeholder="Name" required/>
          <input name="phone" class="field" placeholder="Phone" required/>
        </div>
        <input name="email" class="field" placeholder="Email (optional)"/>
        <div class="grid gap-3 md:grid-cols-2">
          <input name="suburb" class="field" placeholder="Suburb" required/>
          <select name="service" class="field" required>
            <option value="" selected>Select service</option>
            <option>Residential painting</option>
            <option>Commercial painting</option>
            <option>Surface preparation</option>
            <option>Premium finish / repaint</option>
          </select>
        </div>
        <textarea name="message" class="field min-h-[140px]" placeholder="Job details (areas, repairs, timeframe)…" required></textarea>
        <button class="btn btn-primary w-full" type="submit">Send request</button>
        <div id="quoteResult"></div>
        <div class="mt-2 text-xs text-slate-600">Stored locally in leads.sqlite3</div>
      </form>
    </div>
  </div>
</section>
"""
    return HTMLResponse(page("Contact", "/contact", content))


@app.post("/api/quote", response_class=HTMLResponse)
def quote(
    request: Request,
    name: str = Form(...),
    phone: str = Form(...),
    email: Optional[str] = Form(None),
    suburb: str = Form(...),
    service: str = Form(...),
    message: str = Form(...),
    page: str = Form("/"),
):
    name = (name or "").strip()
    phone = (phone or "").strip()
    email = (email or "").strip() or None
    suburb = (suburb or "").strip()
    service = (service or "").strip()
    message = (message or "").strip()
    page = (page or "/").strip()

    errors = []
    if len(name) < 2:
        errors.append("Please enter your name.")
    if len(phone) < 6:
        errors.append("Please enter a valid phone number.")
    if len(suburb) < 2:
        errors.append("Please enter your suburb.")
    if len(service) < 2:
        errors.append("Please choose a service.")
    if len(message) < 5:
        errors.append("Please add a little more detail about the job.")

    if errors:
        items = "".join(f"<li>{e}</li>" for e in errors)
        return HTMLResponse(
            f"""
            <div class="mt-3 rounded-2xl border border-rose-400/20 bg-rose-400/10 p-4 text-rose-200">
              <div class="font-semibold">Please fix:</div>
              <ul class="mt-2 list-disc space-y-1 pl-5 text-sm">{items}</ul>
            </div>
            """
        )

    created_at = datetime.now(timezone.utc).isoformat()
    with sqlite3.connect(DB_PATH) as con:
        con.execute(
            """
            INSERT INTO leads (created_at, name, phone, email, suburb, service, message)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (created_at, name, phone, email, suburb, service, message),
        )
        con.commit()
    lead = {
        "created_at": datetime.utcnow().isoformat(timespec="seconds"),
        "name": name,
        "phone": phone,
        "email": email,
        "state": suburb,
        "service": service,
        "message": message,
        "ip": request.client.host if request.client else "",
    }

    append_lead_to_excel(lead)

    email_part = f" or <span class='font-semibold'>{email}</span>" if email else ""
    return HTMLResponse(
        f"""    
        <div class="mt-5 rounded-2xl p-5 noticeSuccess">
            <div class="title text-lg flex items-center gap-2">
                Request received ✅
            </div>
            <div class="sub mt-1 text-sm">
                We’ll contact <span class="font-semibold">{name}</span> on <span class="font-semibold">{phone}</span> {email_part}.
            </div>
        </div>

        """
    )


@app.get("/robots.txt")
def robots():
    return Response("User-agent: *\nAllow: /\nSitemap: /sitemap.xml\n", media_type="text/plain")


@app.get("/sitemap.xml")
def sitemap():
    pages = ["/", "/services", "/portfolio", "/contact"]
    xml = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">',
        *[f"  <url><loc>{p}</loc></url>" for p in pages],
        "</urlset>",
    ]
    return Response("\n".join(xml), media_type="application/xml")